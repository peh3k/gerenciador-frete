import requests
import json
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import teste
import tkinter as tk
from tkinter import ttk
class DbLink():
    URL_DB = 'https://frete-calculator-default-rtdb.firebaseio.com/'
    PADRAO_DADOS = [
    [
        "ID",
        "nome",
        "peso inicial",
        "peso final",
        "cep inicial",
        "cep final",
        "prazo",
        "estado",
        "cidade",
        "regiao",
        "valor frete",
        "frete min",
        "tac",
        "gris",
        "advalorem",
        "pedagio",
        "tas",
        "icms",
        "outros"
        ], [
            "ID",
            "codigo interno",
            "descricao",
            "unidade",
            "valor venda",
            "peso",
            "comprimento",
            "largura",
            "altura"
        ]
    
    ]

    PATHS = {
        "1": 'Transportadora',
        "2": 'Produto'
    }
    


def post_db(path, dados):
    requisicao = requests.post(f'{DbLink().URL_DB}/{path}/.json', data=json.dumps(dados))


def patch_db(path, dados):
    requisicao = requests.patch(f'{DbLink().URL_DB}/{path}/.json', data=json.dumps(dados))

def get_db(path, id=0, last=False, conflict=False, data=[], names=False, find_name='', find_key_by_name=''):
    requisicao = requests.get(f'{DbLink().URL_DB}/{path}/.json')
    if id != 0:
        items = [item for item in requisicao.json() if requisicao.json()[item]['ID'] == id]

        return items
    if last:
        try:
            items = [item for item in requisicao.json()]

            return requisicao.json()[items[-1]]['ID'] + 1
        except: 
            return 0

    if conflict:
        items = [requisicao.json()[item]['ID'] for item in requisicao.json()]
        same_values = [value[0] for value in data if value[0] in items]
        
        return same_values

    if names:
        empresas = [requisicao.json()[nome]['nome'] for nome in requisicao.json() ]

        return empresas

    if len(find_name) > 0:
        empresa = [requisicao.json()[nome] for nome in requisicao.json() if requisicao.json()[nome]['nome'] == find_name]

        return empresa[0]
    
    if len(find_key_by_name) > 0:
        key = [chave for chave in requisicao.json() if requisicao.json()[chave]['nome'] == find_key_by_name]

        return key[0]

    return requisicao.json()



def find_key_by_code(path, codigo):
    requisicao = requests.get(f'{DbLink().URL_DB}/{path}/.json')
    key = [chave for chave in requisicao.json() if requisicao.json()[chave]['codigo interno'] == codigo]

    return key[0]
def get_excel_rows(excel_file):
    # Ler o arquivo do Excel com pandas
    df = pd.read_excel(excel_file)
    
    df.fillna(value="-", inplace=True)
    rows = df.iloc[0:].values.tolist()
    
    return rows

def delete_db(path):
    requests.delete(f'{DbLink().URL_DB}{path}/.json')


def upload_massivo_db(file, path):
    rows_excel = get_excel_rows(file)
    todos_valores = get_db(path)
    values = [value for value in todos_valores.values()]
    only_ids = [item_[0] for item_interno in values for item_ in item_interno]
    dados_unicos = [dado for dado in rows_excel if dado[0] not in only_ids], [dado for dado in rows_excel if dado[0] in only_ids]

    post_db(path, dados_unicos[0])

    return dados_unicos[1]
    

def dicts_to_lists(dicts):
    # Inicializa a lista de listas
    result = []
    # Pega as chaves do primeiro dicionário para definir a ordem dos valores
    keys = list(dicts[0].keys())
    # Adiciona a lista de chaves na lista de resultados
    result.append(keys)
    # Para cada dicionário na lista de dicionários
    for d in dicts:
        # Cria uma nova lista com os valores do dicionário na ordem das chaves
        values = [d[key] for key in keys]
        # Adiciona a lista de valores na lista de resultados
        result.append(values)

    return result


def criar_tabela_excel(nomes_colunas, linhas, nome_arquivo):
    # Cria um novo arquivo XLSX
    workbook = openpyxl.Workbook()
    # Seleciona a planilha ativa
    sheet = workbook.active
    # Define as colunas com os nomes da lista nomes_colunas
    for coluna, nome_coluna in enumerate(nomes_colunas):
        # A primeira coluna em Excel é "A", portanto, adicionamos 1 ao índice da coluna
        sheet.cell(row=1, column=coluna+1, value=nome_coluna)
    # Define as linhas com os valores da lista linhas
    for linha, valores_linha in enumerate(linhas):
        for coluna, valor in enumerate(valores_linha):
            # A primeira linha em Excel é "1", portanto, adicionamos 2 ao índice da linha
            sheet.cell(row=linha+2, column=coluna+1, value=valor)
    # Salva o arquivo com o nome especificado
    workbook.save(nome_arquivo)

def criar_tabela(frame, colunas, linhas):
    bg_colors = ["white", "gray"]
    frame_label = ttk.Frame(frame)
    frame_tabela = ttk.Frame(frame)
    frame_tabela.pack(fill='both', expand=True)
    style = ttk.Style()
    style.configure("Custom.Treeview", highlightthickness=0,
                    bd=2, relief="groove")
    # cria a lista de colunas e linhas

    # define o objeto Treeview
    tree = ttk.Treeview(frame_tabela, style="Custom.Treeview",
                        columns=colunas, show='headings')

    # define o nome das colunas
    for col in colunas:
        tree.heading(col, text=col)

    for a in linhas:
        tree.insert('', 'end', values=a)

    # configura as colunas para se ajustarem à largura dos dados
    for col in colunas:
        tree.column(col, width=130, anchor='center')

    # verifica se o número de colunas é maior que a largura da janela
    if len(colunas) > 1:
        # cria um scrollbar horizontal
        hsb = ttk.Scrollbar(
            frame_tabela, orient='horizontal', command=tree.xview)
        hsb.pack(side='bottom', fill='x')
        tree.configure(xscrollcommand=hsb.set)

    # cria um scrollbar vertical
    vsb = ttk.Scrollbar(
        frame_tabela, orient='vertical', command=tree.yview)
    vsb.pack(side='right', fill='y')
    tree.configure(yscrollcommand=vsb.set)
    tree.tag_configure("white", background="white")
    tree.tag_configure("gray", background="gray")

    # Aplicar as cores de fundo às linhas da tabela
    for i, item in enumerate(tree.get_children()):
        bg_color = bg_colors[i % len(bg_colors)]
        tree.item(item, tags=(bg_color,))
    # adiciona a tabela e configura o layout
    tree.pack(fill='both', expand=True)




